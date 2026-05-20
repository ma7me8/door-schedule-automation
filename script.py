import openpyxl.worksheet.formula
import openpyxl
from pathlib import Path


HEADER_ROW = 2
DATA_START_ROW = 3
KEY_COLUMN = "ROOM NUMBER"


# this function to read all the doors in all floors
def read_floor_file(file_path: Path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active

    header = {}

    for col in range(1, sheet.max_column + 1):
        header_name = sheet.cell(HEADER_ROW, col).value
        header[header_name] = col

    floor_doors = {}

    # this loop to get all the doors in all floors
    for row in range(DATA_START_ROW, sheet.max_row + 1):
        room_number = sheet.cell(row, header[KEY_COLUMN]).value

        if not room_number:
            continue   # this condition to skip the rows if room number is None

        door_data = {}

        for header_name, col in header.items():    # this loop to get all header names and their values in all floors
            value = sheet.cell(row, col).value
            door_data[header_name] = value

        floor_doors[room_number] = door_data

    return floor_doors


# this function to collect all the doors in all floors
def collect_all_floors(floor_files):
    all_doors = {}

    for floor_file in floor_files:                  # this loop to get all the doors in all floors
        floor_data = read_floor_file(floor_file)
        all_doors.update(floor_data)
    return all_doors


# this function to update the main file
def update_main_file(main_file, all_doors):

    workbook = openpyxl.load_workbook(main_file)
    sheet = workbook.active

    def get_header(sheet):

        header = {}             # this dict to store header names and their values

        # this loop start from first column to the last column in the row
        for col in range(1, sheet.max_column + 1):
            header_name = sheet.cell(HEADER_ROW, col).value

            if header_name is None:
                continue
            header[header_name] = col

        return header

    def update_doors(sheet, header, all_doors):

        update_counter = 0

        for row in range(DATA_START_ROW, sheet.max_row + 1):
            room_number = sheet.cell(row, header["ROOM NUMBER"]).value

            if room_number is None:
                continue

            if not room_number:
                continue

            if room_number not in all_doors:
                continue

            # this dict to get all the doors in all floors
            door_data = all_doors[room_number]

            for header_name, value in door_data.items():

                if header_name not in header:
                    continue
                cell_col = header[header_name]

                sheet.cell(row, cell_col).value = value

            update_counter += 1
        return update_counter

    header = get_header(sheet)
    update_counter = update_doors(sheet, header, all_doors)

    # this line to update the file name
    updated_file = main_file.with_name(
        main_file.stem + "_updated" + main_file.suffix)
    workbook.save(updated_file)

    return update_counter


# main
files = list(Path(r"secondary file path").glob("*.xlsx"))  # type hint

floor_files = [
    file for file in files if not file.name.startswith("~$")
]
print("reading:", floor_files)  # this line to print the file name

all_doors = collect_all_floors(floor_files)

# this line to get the main file name
main_file = Path(r"main file path.xlsx")


updated_count = update_main_file(main_file, all_doors)

print(f"updated {updated_count} doors")
